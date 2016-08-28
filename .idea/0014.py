
'''第 0014 题： 纯文本文件 student.txt为学生信息, 里面的内容（包括花括号）如下所示：
{
    "1":["张三",150,120,100],
    "2":["李四",90,99,95],
    "3":["王五",60,66,68]
}
# 请将上述内容写到 student.xls 文件中。'''

import json
from openpyxl import Workbook
def txt2xls(path):
    file = open(path,'r',encoding='UTF-8')
    content =''
    for line in file.readlines():
        if line.startswith(u'\ufeff'):
            line = line.encode('utf8')[3:].decode('utf8')
        content = content+line

    jsonResult = json.loads(content,encoding='utf-8')
    print(content)
    print(jsonResult)

    workbook = Workbook();
    workSheet = workbook.worksheets[0]

    length1 = len(jsonResult)
    print(len(jsonResult))
    for i in range(1, len(jsonResult) + 1):
        workSheet.cell(row=i, column=1).value = i
        print(jsonResult[str(i)])
        length2 = len(jsonResult[str(i)])
        for m in range(0, len(jsonResult[str(i)])):
            workSheet.cell(row=i, column=m + 2).value = jsonResult[str(i)][m]


    workbook.save("student.xls")

txt2xls('student.txt')