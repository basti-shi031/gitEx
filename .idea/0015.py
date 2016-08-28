import json
from openpyxl import Workbook

def txt2xls(path):
    content = openfile(path)
    contentJson = json.loads(content, encoding='utf-8')
    print(contentJson)
    json2xls(contentJson)


def json2xls(contentJson):
    workBook = Workbook()
    workSheet = workBook.worksheets[0];
    size = len(contentJson)
    for i in range(size):
        workSheet.cell(column=1,row=i+1).value = i+1
        contentB = contentJson[str(i+1)]
        workSheet.cell(column=2,row=i+1).value = contentB
    workBook.save("city.xls")
def openfile(path):
    file = open(path, encoding='utf-8')
    content = ''
    for line in file.readlines():
        if line.startswith(u'\ufeff'):
            line = line.encode('utf8')[3:].decode('utf8')
        content = content + line
    return content


txt2xls('city.txt');
